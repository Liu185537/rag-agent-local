from __future__ import annotations

import csv
import io
import json
from pathlib import Path


def parse_document(file_name: str, content_bytes: bytes) -> str:
    """根据文件后缀选择解析策略并返回纯文本。"""
    suffix = Path(file_name).suffix.lower()
    if suffix in {".txt", ".md", ".log"}:
        return content_bytes.decode("utf-8", errors="ignore")
    if suffix == ".csv":
        return _parse_csv(content_bytes)
    if suffix == ".json":
        return _parse_json(content_bytes)
    if suffix == ".pdf":
        return _parse_pdf(content_bytes)
    if suffix == ".docx":
        return _parse_docx(content_bytes)
    if suffix == ".xlsx":
        return _parse_xlsx(content_bytes)
    raise ValueError(f"Unsupported file type: {suffix}")


def _parse_csv(content_bytes: bytes) -> str:
    """CSV 转文本：每行转为以 | 分隔的可读字符串。"""
    text = content_bytes.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    lines = []
    for row in reader:
        if not row:
            continue
        lines.append(" | ".join(col.strip() for col in row))
    return "\n".join(lines)


def _parse_json(content_bytes: bytes) -> str:
    """JSON 转文本：优先格式化，失败则返回原文。"""
    text = content_bytes.decode("utf-8", errors="ignore")
    try:
        obj = json.loads(text)
        return json.dumps(obj, ensure_ascii=False, indent=2)
    except Exception:
        return text


def _parse_pdf(content_bytes: bytes) -> str:
    """PDF 转文本：逐页提取后拼接。"""
    try:
        from pypdf import PdfReader
    except Exception as exc:
        raise RuntimeError("pypdf is required for PDF parsing. Install with: pip install pypdf") from exc

    reader = PdfReader(io.BytesIO(content_bytes))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(pages).strip()


def _parse_docx(content_bytes: bytes) -> str:
    """DOCX 转文本：抽取非空段落。"""
    try:
        import docx
    except Exception as exc:
        raise RuntimeError(
            "python-docx is required for DOCX parsing. Install with: pip install python-docx"
        ) from exc

    document = docx.Document(io.BytesIO(content_bytes))
    lines = [p.text.strip() for p in document.paragraphs if p.text and p.text.strip()]
    return "\n".join(lines)


def _parse_xlsx(content_bytes: bytes) -> str:
    """XLSX 转文本：按 sheet 和行展开为可读文本。"""
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required for XLSX parsing. Install with: pip install openpyxl") from exc

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in wb.worksheets:
        lines.append(f"# sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if cells:
                lines.append(" | ".join(cells))

    return "\n".join(lines).strip()

